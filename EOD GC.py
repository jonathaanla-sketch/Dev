import time

from datetime import datetime, timedelta

from selenium import webdriver

from selenium.webdriver.chrome.service import Service

from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.common.keys import Keys

from webdriver_manager.chrome import ChromeDriverManager


import os

import glob

import pandas as pd

import openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.utils import get_column_letter


# --- 1. CONFIGURACIÓN DE ACCESO ---

USER = "Jonathanlopez@contento-goodcharlie.gocontact.com"

PASS = "Noviembre2023*"

URL_LOGIN = "https://contento-goodcharlie.gocontact.com/"

URL_REPORTE = "https://contento-goodcharlie.gocontact.com/#/fs/modules/report-builder/report-builder.html"



# --- 2. LÓGICA DE FECHAS (Lunes -> Viernes / Resto -> Ayer) ---

hoy = datetime.now()

if hoy.weekday() == 0:  # 0 es Lunes

    fecha_objetivo = hoy - timedelta(days=3)

else:

    fecha_objetivo = hoy - timedelta(days=1)



fecha_str = fecha_objetivo.strftime("%Y-%m-%d")



# --- 3. INICIALIZACIÓN DEL NAVEGADOR ---

options = webdriver.ChromeOptions()

options.add_experimental_option("detach", True)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

wait = WebDriverWait(driver, 30)



try:

    # --- PASO A: LOGIN REFORZADO ---

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Iniciando sesión...")

    driver.get(URL_LOGIN)

   

    u_field = wait.until(EC.element_to_be_clickable((By.NAME, "email")))

    u_field.send_keys(USER)



    try:

        p_field = driver.find_element(By.NAME, "password_label")

    except:

        p_field = driver.find_element(By.CSS_SELECTOR, "input[type='password']")



    p_field.click()

    p_field.clear()

    for letra in PASS:

        p_field.send_keys(letra)

        time.sleep(0.05)

    p_field.send_keys(Keys.ENTER)

   

    # --- PASO B: NAVEGACIÓN ---

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Login ok. Entrando a reportes...")

    time.sleep(10)

    driver.get(URL_REPORTE)

    # Espera extendida para asegurar que los scripts de la página (jQuery/Select2) carguen

    time.sleep(10)



    # --- PASO C: FILTROS (INYECCIÓN JS CON TIEMPOS DE ASENTAMIENTO) ---

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Aplicando filtros...")



    # 1. Data Source -> 'Queues' (valor: 'queue')

    driver.execute_script("$('#dataSource').val('queue').trigger('change');")

    print("DataSource seleccionado (Queues). Esperando carga de templates...")

    time.sleep(4) # Tiempo vital para que la lista de templates se actualice



    # 2. Template -> 'Good Charlie Reporting Template' (valor: '27')

    driver.execute_script("$('#selectTemplate').val('27').trigger('change');")

    time.sleep(2)



    # 3. Type of Data -> 'All calls' (valor: '0')

    driver.execute_script("$('#dataTypeCalls').val('0').trigger('change');")

    time.sleep(2)



    # 4. Fechas (Start y End)

    print(f"Asignando fechas: {fecha_str}")

    script_fechas = f"""

        $('#startDate').val('{fecha_str}').trigger('change');

        $('#endDate').val('{fecha_str}').trigger('change');

    """

    driver.execute_script(script_fechas)

    time.sleep(3) # Espera final para que la validación de la página reconozca los datos



    # --- PASO D: GENERACIÓN DEL REPORTE (CON REFUERZO DE VALIDACIÓN) ---
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Sincronizando validaciones y descargando...")

    # 4. Fechas (Start y End)

    print(f"Asignando fechas: {fecha_str}")

    script_fechas = f"""

        $('#startDate').val('{fecha_str}').trigger('change');

        $('#endDate').val('{fecha_str}').trigger('change');

    """

    driver.execute_script(script_fechas)

    time.sleep(3) # Espera final para que la validación de la página reconozca los datos

    # Localizamos el botón por su ID final
    btn_final = wait.until(EC.element_to_be_clickable((By.ID, "downloadreportsubmit")))
    
    # Scroll preventivo
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn_final)
    time.sleep(1)
    
    # CLIC DEFINITIVO
    driver.execute_script("arguments[0].click();", btn_final)

    # Tiempo de descarga solicitado
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Esperando 70 segundos para la descarga...")
    time.sleep(5)

   

    print(f"✅ ¡Proceso completado! Archivo para la fecha {fecha_str} generado.")



except Exception as e:

    print(f"\n❌ Se produjo un error: {e}")

    driver.save_screenshot("error_reporte_gocontact.png")

    print("Se ha guardado una captura de pantalla como 'error_reporte_gocontact.png'.")



# driver.quit() # Mantener para verificar



# Ruta de Descargas

downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")

csv_files = glob.glob(os.path.join(downloads_path, "*.csv"))



if not csv_files:

    print("No se encontraron archivos CSV en Descargas.")

else:

    latest_csv = max(csv_files, key=os.path.getctime)

    print("Archivo seleccionado:", os.path.basename(latest_csv))



    # Leer el archivo con separador de punto y coma

    df = pd.read_csv(latest_csv, encoding="utf-8", sep=";")

    columnas = df.columns.tolist()



    # ---- Filtros ----

    filtro1 = (df.iloc[:, 1] == 0) & (df.iloc[:, 42] != "ABANDON")

    filtro2 = (df.iloc[:, 23] == "7867362870")

    filtro3 = df.iloc[:, 35].astype(str).str.lower().str.contains("coll", na=False)



    filtro_total = filtro1 | filtro2 | filtro3

    df_filtrado = df.drop(df[filtro_total].index)



    # ---- UTM ----

    utm_valores = ["Digital Discount", "PTC Standard", "Digital Standard", "PTC Discount"]

    utm_rows = df_filtrado[df_filtrado.iloc[:, 35].isin(utm_valores)]

    utm_volume = utm_rows.shape[0]



    def aplicar_formatos(df):

        df = df.copy()

        df[columnas[13]] = pd.to_datetime(df[columnas[13]], errors="coerce")

        df[columnas[18]] = pd.to_datetime(df[columnas[18]], errors="coerce")

        df[columnas[23]] = pd.to_numeric(df[columnas[23]], errors="coerce")

        return df



    fecha_str = "unknown"



    # ---- Guardar archivo UTM ----

    if utm_volume > 0:

        utm_rows_fmt = aplicar_formatos(utm_rows)

        fecha = pd.to_datetime(utm_rows_fmt.iloc[0, 18], errors="coerce")

        if pd.notnull(fecha):

            fecha_str = fecha.strftime("%m-%d")



        filename = f"UTM EOD {fecha_str}.xlsx"

        filepath = os.path.join(downloads_path, filename)



        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

            utm_rows_fmt.to_excel(writer, sheet_name="UTM Data", index=False)



            # ---- Conteo por Queue ----

            conteo_queue = utm_rows_fmt.iloc[:, 35].value_counts().reset_index()

            conteo_queue.columns = ["Queue", "Count"]

            conteo_queue["Percentage"] = conteo_queue["Count"] / conteo_queue["Count"].sum()



            total_count = conteo_queue["Count"].sum()

            conteo_queue.loc[len(conteo_queue)] = ["Total", total_count, 1.0]



            conteo_queue.to_excel(writer, sheet_name="Queue Summary", index=False)



        wb = openpyxl.load_workbook(filepath)

        ws = wb["Queue Summary"]



        end_row = ws.max_row

        table_ref = f"A1:C{end_row}"



        table = Table(displayName="QueueTable", ref=table_ref)

        style = TableStyleInfo(name="TableStyleMedium1", showRowStripes=True, showColumnStripes=False)

        table.tableStyleInfo = style

        ws.add_table(table)



        # Formato de porcentaje en la columna C

        for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=3, max_col=3):

            for cell in row:

                cell.number_format = "0.00%"



        # Negrilla en la fila de Total

        for cell in ws[end_row]:

            cell.font = openpyxl.styles.Font(bold=True)



        # Autoajustar columnas

        for col in ws.columns:

            max_length = 0

            col_letter = col[0].column_letter

            for cell in col:

                try:

                    if cell.value:

                        max_length = max(max_length, len(str(cell.value)))

                except:

                    pass

            ws.column_dimensions[col_letter].width = max_length + 2



        wb.active = wb["Queue Summary"]

        wb.save(filepath)

        os.startfile(filepath)



        print(f"Archivo UTM guardado y abierto en: {filepath}")

    else:

        print("No se encontraron filas UTM para guardar.")



    # ---- Guardar archivo de llamadas reales ----

    if len(df_filtrado) > 0:

        df_llamadas_fmt = aplicar_formatos(df_filtrado)

        fecha_llamadas = pd.to_datetime(df_llamadas_fmt.iloc[0, 18], errors="coerce")

        if pd.notnull(fecha_llamadas):

            fecha_llamadas_str = fecha_llamadas.strftime("%m%d")

        else:

            fecha_llamadas_str = "unknown"



        filename_llamadas = f"llamadas reales {fecha_llamadas_str}.xlsx"

        filepath_llamadas = os.path.join(downloads_path, filename_llamadas)



        with pd.ExcelWriter(filepath_llamadas, engine="openpyxl") as writer:

            df_llamadas_fmt.to_excel(writer, sheet_name="Llamadas Reales", index=False)



        wb_ll = openpyxl.load_workbook(filepath_llamadas)

        ws_ll = wb_ll["Llamadas Reales"]



        end_row_ll = ws_ll.max_row

        end_col_ll = ws_ll.max_column

        table_ref_ll = f"A1:{get_column_letter(end_col_ll)}{end_row_ll}"



        table_ll = Table(displayName="LlamadasTable", ref=table_ref_ll)

        style_ll = TableStyleInfo(name="TableStyleLight11", showRowStripes=True, showColumnStripes=False)

        table_ll.tableStyleInfo = style_ll

        ws_ll.add_table(table_ll)



        # Autoajustar columnas

        for col in ws_ll.columns:

            max_length = 0

            col_letter = col[0].column_letter

            for cell in col:

                try:

                    if cell.value:

                        max_length = max(max_length, len(str(cell.value)))

                except:

                    pass

            ws_ll.column_dimensions[col_letter].width = max_length + 2



        wb_ll.active = wb_ll["Llamadas Reales"]

        wb_ll.save(filepath_llamadas)

        os.startfile(filepath_llamadas)



        print(f"Archivo de llamadas reales guardado y abierto en: {filepath_llamadas}")

    else:

        print("No hay llamadas finales para guardar.")
